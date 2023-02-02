import tkinter
import tkinter.messagebox
import customtkinter
import pandas as pd
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from time import sleep
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows

###################### FUNCTIONS #######################
########################################################


rankings_criteria_params = {}
rankings_criteria_params["arwu_criteria_params"] = {1: "Alumni", 2: "Award", 3: "HiCi", 4: "N&S", 5: "PUB", 6: "PCP"}
rankings_criteria_params["grsssd_criteria_params"] = {1: "PUB", 2: "CIT", 3: "CPP", 4: "TOP", 5: "IC"}

def sub_scrap_shangai_table(browser,ranking, criteria_label, result):
    soup = BeautifulSoup(browser.page_source, 'lxml')
    for row in soup.select("table.rk-table > tbody > tr"):
        temp_dict = {}
        world_ranking = row.find("div", class_= re.compile('^ranking'))
        univ_name = row.find("span", class_="univ-name")
        country = row.find("div", class_="region-img", style=True)
        if ranking == "arwu":
            total_score = row.findAll('td')[4]
            param_score = row.findAll('td')[5]
        if ranking == "grsssd":
            total_score = row.findAll('td')[3]
            param_score = row.findAll('td')[4]
        temp_dict[f"{ranking}_world_rank"] = world_ranking.text.strip()
        temp_dict["univ_name"] = univ_name.text.strip()
        temp_dict["country"] = country["style"].partition('png100/')[2].partition(".png")[0]
        temp_dict[f"{ranking}_total_score"] = total_score.text.strip()
        temp_dict[f"{criteria_label}_score"] = param_score.text.strip()
        result.append(temp_dict)

def get_shangai_data(tk_container,year=None, ranking=None, max_pages=None):
    browser = webdriver.Firefox()
    browser.set_window_position(0, 0)
    browser.set_window_size(200, 200)
    try:
        browser.get(f'https://www.shanghairanking.com/rankings/{ranking}/{year}')
    except:
        tk_container.insert(tkinter.END, "Error, check if the selected ranking exists for the selected year\n\n")   
        app.update()  
    finally:   
        sleep(2)
    cols_number = len(rankings_criteria_params[f"{ranking}_criteria_params"])
    df_dict = {}
    if max_pages is None:
        max_pages = browser.find_element(by = By.XPATH, value = "//ul[@class='ant-pagination']/li[last()-1]/a").text
    for item_key,item_value in rankings_criteria_params[f"{ranking}_criteria_params"].items():
        tk_container.insert(tkinter.END, "Scraping launched for "+ item_value + "...\n\n")
        app.update()
        #browser.find_element(by = By.XPATH, value = "//table[@class='rk-table']/thead/tr[1]/th["+str(cols_number)+"]/div/div/div/input[@class='head-bg']").click()  
        #browser.find_element(by = By.XPATH, value = "//table[@class='rk-table']/thead/tr[1]/th["+str(cols_number)+"]/div/div/div[@class='rk-tooltip']/ul/li["+str(item_key)+"]").click()
        el_head = browser.find_element(by = By.XPATH, value = "//table[@class='rk-table']/thead/tr[1]/th["+str(cols_number)+"]/div/div/div/input[@class='head-bg']")
        browser.execute_script("arguments[0].click();", el_head)
        el_li = browser.find_element(by = By.XPATH, value = "//table[@class='rk-table']/thead/tr[1]/th["+str(cols_number)+"]/div/div/div[@class='rk-tooltip']/ul/li["+str(item_key)+"]")
        browser.execute_script("arguments[0].click();", el_li)
        result = []
        for i in range(0,int(max_pages)):
            #browser.find_element(by = By.XPATH, value = "//li[@title='"+str(i+1)+"']").click()
            el = browser.find_element(by = By.XPATH, value = "//li[@title='"+str(i+1)+"']")
            browser.execute_script("arguments[0].click();", el)
            sleep(1)
            sub_scrap_shangai_table(browser,ranking,item_value,result)
        df_dict[f"df_{year}_{ranking}_{item_value}"] = pd.DataFrame(result)
        df_dict[f"df_{year}_{ranking}_{item_value}"]["year"] = year
        # shift column 'year' to first position
        first_column = df_dict[f"df_{year}_{ranking}_{item_value}"].pop('year')
        df_dict[f"df_{year}_{ranking}_{item_value}"].insert(0, 'year', first_column)
        # make sure the criteria score is number format
        df_dict[f"df_{year}_{ranking}_{item_value}"][f"{item_value}_score"] = pd.to_numeric(df_dict[f"df_{year}_{ranking}_{item_value}"][f"{item_value}_score"], downcast='integer', errors='coerce')
        # add global rank on the argument criteria
        df_dict[f"df_{year}_{ranking}_{item_value}"][f"calculated_{item_value}_world_rank"] = df_dict[f"df_{year}_{ranking}_{item_value}"][f"{item_value}_score"].rank(method="dense", ascending=False)
        # add national rank on the argument criteria
        df_dict[f"df_{year}_{ranking}_{item_value}"][f"calculated_{item_value}_national_rank"] = df_dict[f"df_{year}_{ranking}_{item_value}"].groupby(["country"])[f"{item_value}_score"].rank(method="dense", ascending=False)
        # reset index
        df_dict[f"df_{year}_{ranking}_{item_value}"] = df_dict[f"df_{year}_{ranking}_{item_value}"].reset_index()
    # close selenium driver
    browser.close()
    tk_container.insert(tkinter.END, "End data collecting...\n\n")
    app.update()
    return df_dict

def sub_arwu_score_formula(x):   
    return (x["Alumni_score"]*0.1) + (x["Award_score"]*0.2) + (x["HiCi_score"]*0.2) + (x["N&S_score"]*0.2) + (x["PUB_score"]*0.2) + (x["PCP_score"]*0.1)
def sub_grsssd_score_formula(x):
    return (x["PUB_score"]*0.2) + (x["CIT_score"]*0.2) + (x["CPP_score"]*0.3) + (x["TOP_score"]*0.2) + (x["IC_score"]*0.1)
def sub_calculate_world_score_and_rank(df_result, ranking):
    # on calcule le coefficient d'ajustement (à appliquer ensuite aux scores pondérés de chaque univ) sur la base d'univ n°1
    first = df_result.head(1)
    #first = df_result[df_result[f"{ranking}_total_score"] == '100.0'].head(1)
    #first = df_result.sort_values(by=f'{ranking}_total_score', ascending=False).head(1)
    df_result.to_csv("result.csv",index=False,encoding="utf-8")
    #first = df_result.iloc[[0]]
    print(first)
    if ranking == "arwu":
        coeff = 100 / sub_arwu_score_formula(first)
        df_result["calculated_total_score"] = df_result.apply(lambda x: round(coeff * sub_arwu_score_formula(x),1), axis=1)
    if ranking == "grsssd":
        coeff = 100 / sub_grsssd_score_formula(first)
        df_result["calculated_total_score"] = df_result.apply(lambda x: round(coeff * sub_grsssd_score_formula(x),1), axis=1)
    # make sure the score is number format
    df_result["calculated_total_score"] = pd.to_numeric(df_result["calculated_total_score"], downcast='integer', errors='coerce')
    df_result["calculated_world_rank"] = df_result["calculated_total_score"].rank(method="dense", ascending=False)
    df_result["calculated_national_rank"] = df_result.groupby(["country"])["calculated_total_score"].rank(method="dense", ascending=False)
    return df_result

def calculate_each_criteria_data(df_dict, year, ranking):
    new_df_dict = {}
    for item_value in rankings_criteria_params[f"{ranking}_criteria_params"].values():
        new_df_dict[f"df_{year}_{ranking}_{item_value}"] = (df_dict[f"df_{year}_{ranking}_{item_value}"]
         .drop(columns=[f"{ranking}_world_rank",f"{ranking}_total_score"])
         .sort_values(by=f'{item_value}_score', ascending=False)
         .reset_index()
         .drop(columns=['level_0'])
         )
        new_df_dict[f"df_{year}_{ranking}_{item_value}"]["index"] = new_df_dict[f"df_{year}_{ranking}_{item_value}"]["index"] + 1   
    return new_df_dict

def calculate_main_data(df_dict, year, ranking):
        # pour la boucle
        criteria_length = len(rankings_criteria_params[f"{ranking}_criteria_params"])
        temp = {}
        # on récupère le 1er dataframe du 1 critère tel quel car il contient les données globales du classement
        temp["1"] = df_dict[f"df_{year}_{ranking}_"+rankings_criteria_params[f"{ranking}_criteria_params"][1]]
        # puis on loop en ne conservant que les colonnes calculées des autres critères
        for i in range(1,criteria_length):
            temp[str(i+1)] = pd.merge(temp[str(i)], df_dict[f"df_{year}_{ranking}_"+rankings_criteria_params[f"{ranking}_criteria_params"][i+1]][['univ_name', rankings_criteria_params[f"{ranking}_criteria_params"][i+1]+"_score", "calculated_"+rankings_criteria_params[f"{ranking}_criteria_params"][i+1]+"_world_rank","calculated_"+rankings_criteria_params[f"{ranking}_criteria_params"][i+1]+"_national_rank"]], how='inner', on=['univ_name'])
        df_all = sub_calculate_world_score_and_rank(temp[str(criteria_length)], ranking)
        df_result = (df_all
         .sort_values(by='calculated_total_score', ascending=False)
         .reset_index()
         .drop(columns=['level_0'])
         )
        df_result["index"] = df_result["index"] + 1
        return df_result

def design_worksheet(ws):
    # lits of column names starting with "calculated" 
    list_cols_names=[]
    for cell in ws[1]:
        list_cols_names.append(cell.value)
    # iterate all cols and fill them green
    list_calculated_cols_names = list(filter(lambda x: str(x)[0:10] == "calculated", list_cols_names))
    for column_cell in ws.iter_cols(1, ws.max_column):
        ws.column_dimensions[str(column_cell[0].column_letter)].width = 20
        if str(column_cell[0].value) in list_calculated_cols_names:
            ws.column_dimensions[str(column_cell[0].column_letter)].width = 30
            for row in column_cell[0:]:
                row.fill = PatternFill("solid", start_color="5cb800")
    
def create_workbook(tk_container, df_dict,year, ranking):
    tk_container.insert(tkinter.END, "Design Excel results file...\n\n")
    app.update()
    wb = Workbook()
    ## compile & insert & design all data
    ws_all = wb.active
    ws_all.title = "All"
    all_data = calculate_main_data(df_dict, year, ranking)
    for r in dataframe_to_rows(all_data, index=False, header=True):
         ws_all.append(r)
    design_worksheet(ws_all)
    ## compile & insert & design all data
    each_criteria_data = calculate_each_criteria_data(df_dict, year, ranking)
    for item_value in rankings_criteria_params[f"{ranking}_criteria_params"].values():
        ws = wb.create_sheet(f'{item_value}')
        for r in dataframe_to_rows(each_criteria_data[f"df_{year}_{ranking}_{item_value}"], index=False, header=True):
            ws.append(r)
        design_worksheet(ws)
    tk_container.insert(tkinter.END, "Done ! File ready to be saved\n\n")
    # save
    #wb.save("C:/Users/geoffroy/JupyterNotebooks/ranking_scrapping/tkinter.xlsx")
    return wb

#global wb
#wb = Workbook()

#################### UI ############################
####################################################

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

year_values = ["2022", "2021", "2020", "2019", "2018", "2017", "2016", "2015", "2014", "2013"]

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Academic Ranking Scraper")
        self.geometry(f"{900}x{520}+{250}+{200}")

        # configure grid layout (2x2)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self, width=180, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=3, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.ranking_label = customtkinter.CTkLabel(self.sidebar_frame, text="Choisir un ranking", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.ranking_label.grid(row=0, column=0, padx=10, pady=(5, 0))

        # create tabview in sidebar
        self.tabview = customtkinter.CTkTabview(self.sidebar_frame, width=250)
        self.tabview.grid(row=1, column=0, padx=(20, 0), pady=(10, 0), sticky="nsew")
        self.tabview.add("Shangaï")
        self.tabview.add("THE")
        self.tabview.tab("Shangaï").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
        self.tabview.tab("THE").grid_columnconfigure(0, weight=1)

        # Tabview content
        ## Shangaï
        self.ranking_selection = tkinter.StringVar(value="arwu")
        self.radio_button_arwu = customtkinter.CTkRadioButton(self.tabview.tab("Shangaï"), variable=self.ranking_selection, value="arwu", command=self.ranking_selection_event)
        self.radio_button_arwu.grid(row=1, column=0, pady=5, padx=10, sticky="n")
        self.radio_button_grsssd = customtkinter.CTkRadioButton(self.tabview.tab("Shangaï"), variable=self.ranking_selection, value="grsssd", command=self.ranking_selection_event)
        self.radio_button_grsssd.grid(row=2, column=0, pady=5, padx=10, sticky="n")
        ## THE
        self.coming_soon_label = customtkinter.CTkLabel(self.tabview.tab("THE"), text="Coming soon", anchor="center")
        
        # year dropdown
        self.year_label = customtkinter.CTkLabel(self.sidebar_frame, text="Choisir une année", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.year_label.grid(row=2, column=0, padx=10, pady=(5, 0))
        self.year_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=year_values, command=self.change_year_event)
        self.year_mode_optionemenu.grid(row=3, column=0, padx=5, pady=(10, 0))

        # Action buttons
        self.sidebar_button_scrap_pageone = customtkinter.CTkButton(self.sidebar_frame, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), command=self.sidebar_button_scraponepage_event)
        self.sidebar_button_scrap_pageone.grid(row=4, column=0, padx=20, pady=(15,0))
        self.sidebar_button_scrap_all = customtkinter.CTkButton(self.sidebar_frame, fg_color="#4EC687", text_color="#040504", command=self.sidebar_button_scrapall_event)
        self.sidebar_button_scrap_all.grid(row=5, column=0, padx=5, pady=(2,10))
        # create textbox
        self.textbox = customtkinter.CTkTextbox(self, width=650, fg_color="#040504", text_color="#E9F3E9")
        self.textbox.grid(row=0, column=1, rowspan=2, padx=(20, 20), pady=(5, 0), sticky="nsew")

        # clear button
        #self.button_clear = customtkinter.CTkButton(self, text="Effacer", fg_color="#FA655C", border_width=2, text_color="#040504", command=self.button_clear_event)
        self.button_clear = customtkinter.CTkButton(self, text="Réinitialiser", fg_color="transparent", border_width=2, text_color="#FA655C", command=self.button_clear_event)
        self.button_clear.grid(row=1, column=1, padx=20, pady=(350,5), sticky="n")
        
        # Download button
        self.download_excel_result = customtkinter.CTkButton(self, text="Télécharger le fichier de résultats (Excel)", fg_color="#4EC687", text_color="#040504",
                                                           command=self.download_file)

        self.download_excel_result.grid(row=2, column=1, padx=20, pady=(10, 10))

        # Credits 
        self.footer_frame = customtkinter.CTkFrame(self, width=1100, height=50)
        self.footer_frame.grid(row=6, column=0, columnspan=4, sticky="nsew")
        self.footer = customtkinter.CTkLabel(self.footer_frame, text="Powered by D2P Unversité Côte d'Azur")
        self.footer.grid(padx=300, pady=(10, 0), sticky="nsew")

        # set default values
        self.radio_button_arwu.configure(text="Classement ARWU")
        self.radio_button_grsssd.configure(text="Classement GRSSSD")
        self.sidebar_button_scrap_pageone.configure(text="[Debug] Scraper les 2 premières pages")
        self.sidebar_button_scrap_all.configure(text="Scraper toutes les données")
        self.year_mode_optionemenu.set("2022")
        self.textbox.insert("0.0", "Ready to start\n\n")
        self.wb = Workbook()

    def ranking_selection_event(self):
        print(self.ranking_selection.get())

    def change_year_event(self, selected_year: str):
        print(selected_year)
        self.year_mode_optionemenu.set(selected_year)

    def sidebar_button_scraponepage_event(self):
        year = self.year_mode_optionemenu.get()
        ranking = self.ranking_selection.get()
        self.textbox.insert(tkinter.END, "Scraping launched...\n\n")
        df = get_shangai_data(self.textbox, year=year, ranking=ranking, max_pages=2)
        self.textbox.insert(tkinter.END, "Save and design in Excel...\n\n")
        self.wb = create_workbook(self.textbox, df,year, ranking)

    def sidebar_button_scrapall_event(self):
        year = self.year_mode_optionemenu.get()
        ranking = self.ranking_selection.get()
        self.textbox.insert(tkinter.END, "Scraping launched...\n\n")
        df = get_shangai_data(self.textbox, year=year, ranking=ranking)
        self.textbox.insert(tkinter.END, "Save and design in Excel...\n\n")
        self.wb = create_workbook(self.textbox, df,year, ranking)

    def download_file(self):
        file_path = tkinter.filedialog.asksaveasfilename(filetypes = [('All types(*.*)', '*.*'),("Excel file(*.xlsx)","*.xlsx")], defaultextension = [('All types(*.*)', '*.*'),("Excel file(*.xlsx)","*.xslx")])
        if file_path:
            self.wb.save(file_path)

    def button_clear_event(self):
        self.textbox.delete("1.0",tkinter.END)
        self.textbox.insert("0.0", "Ready to start\n\n")
        self.wb = Workbook()


if __name__ == "__main__":
    app = App()
    app.mainloop()